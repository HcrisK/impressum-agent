import streamlit as st
import anthropic
import requests
import json
import re
import openpyxl
import io
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import httpx
import time

# ── Keys aus Streamlit Secrets ──────────────────
ANTHROPIC_KEY = st.secrets["ANTHROPIC_KEY"]
HUBSPOT_KEY   = st.secrets["HUBSPOT_KEY"]

client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

# ── Impressum finden (ohne Selenium!) ──────────
def basis_url(url):
    p = urlparse(url)
    return f"{p.scheme}://{p.netloc}"

def impressum_url_finden(url):
    basis = basis_url(url)
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
    try:
        r = httpx.get(basis, headers=headers, follow_redirects=True, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        for link in soup.find_all("a", href=True):
            text = link.get_text(strip=True).lower()
            href = link["href"].lower()
            if any(k in text or k in href for k in ["impressum", "imprint", "legal"]):
                from urllib.parse import urljoin
                return urljoin(basis, link["href"])
    except:
        pass
    for pfad in ["/impressum", "/imprint", "/legal"]:
        versuch = basis + pfad
        try:
            r = httpx.get(versuch, headers=headers, timeout=5)
            if r.status_code == 200:
                return versuch
        except:
            pass
    return basis

def seite_laden(url):
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
    try:
        r = httpx.get(url, headers=headers, follow_redirects=True, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "header"]):
            tag.decompose()
        zeilen = [z for z in soup.get_text(separator="\n").splitlines() if z.strip()]
        return "\n".join(zeilen)[:5000]
    except:
        return ""

def claude_extrahieren(text):
    antwort = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": f"""Extrahiere die Impressum-Daten aus dem Text.
Antworte NUR mit einem JSON-Objekt, kein Text davor oder danach, keine Backticks.

{{
  "firmenname": "...",
  "adresse": "...",
  "telefon": "...",
  "email": "...",
  "geschaeftsfuehrer": "...",
  "webseite": "..."
}}

Felder die nicht vorhanden sind mit null befüllen.

Text:
{text}"""
        }]
    )
    rohtext = antwort.content[0].text
    match = re.search(r'\{[\s\S]*\}', rohtext)
    if match:
        return json.loads(match.group())
    return {}

def hubspot_kontakt_anlegen(daten, quell_url):
    url = "https://api.hubapi.com/crm/v3/objects/contacts"
    headers = {
        "Authorization": f"Bearer {HUBSPOT_KEY}",
        "Content-Type": "application/json"
    }
    name_teile = (daten.get("geschaeftsfuehrer") or "").split(" ", 1)
    vorname = name_teile[0] if len(name_teile) > 0 else ""
    nachname = name_teile[1] if len(name_teile) > 1 else ""
    eigenschaften = {
        "company":        daten.get("firmenname") or "",
        "address":        daten.get("adresse") or "",
        "phone":          daten.get("telefon") or "",
        "email":          daten.get("email") or "",
        "firstname":      vorname,
        "lastname":       nachname,
        "website":        quell_url,
        "hs_lead_status": "NEW",
    }
    eigenschaften = {k: v for k, v in eigenschaften.items() if v}
    antwort = requests.post(url, headers=headers, json={"properties": eigenschaften})
    if antwort.status_code == 201:
        return "✅ Kontakt angelegt"
    elif antwort.status_code == 409:
        return "ℹ️ Kontakt existiert bereits"
    else:
        return f"❌ Fehler: {antwort.text}"

# ── Streamlit UI ────────────────────────────────
st.set_page_config(page_title="Impressum Agent", page_icon="🔍")
st.title("🔍 Impressum Agent")
st.write("Excel-Datei mit URLs hochladen — der Agent extrahiert automatisch alle Impressum-Daten und trägt sie in HubSpot ein.")

datei = st.file_uploader("Excel-Datei hochladen", type=["xlsx"])

if datei and st.button("▶ Starten"):
    wb = openpyxl.load_workbook(io.BytesIO(datei.read()))
    ws = wb.active
    urls = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]

    st.info(f"{len(urls)} URLs gefunden. Starte...")
    fortschritt = st.progress(0)
    status_text = st.empty()
    log = st.empty()

    ergebnisse = []
    log_zeilen = []

    for i, url in enumerate(urls, 1):
        status_text.text(f"Verarbeite {i} von {len(urls)}: {url}")
        try:
            imp_url = impressum_url_finden(url)
            text = seite_laden(imp_url)
            daten = claude_extrahieren(text)
            hs_status = hubspot_kontakt_anlegen(daten, url)
            log_zeilen.append(f"✅ [{i}] {url} → {hs_status}")
            ergebnisse.append({"url": url, "daten": daten, "status": hs_status})
        except Exception as e:
            log_zeilen.append(f"❌ [{i}] {url} → Fehler: {e}")
            ergebnisse.append({"url": url, "daten": {}, "status": f"Fehler: {e}"})

        fortschritt.progress(i / len(urls))
        log.text("\n".join(log_zeilen[-10:]))
        time.sleep(1)

    # Ergebnis-Excel erstellen
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["URL", "Firmenname", "Adresse", "Telefon", "Email", "Geschäftsführer", "Status"])
    for e in ergebnisse:
        d = e.get("daten", {})
        ws2.append([
            e.get("url", ""),
            d.get("firmenname", ""),
            d.get("adresse", ""),
            d.get("telefon", ""),
            d.get("email", ""),
            d.get("geschaeftsfuehrer", ""),
            e.get("status", "")
        ])

    output = io.BytesIO()
    wb2.save(output)
    output.seek(0)

    st.success("✅ Fertig!")
    st.download_button(
        label="📥 Ergebnisse herunterladen",
        data=output,
        file_name="ergebnisse.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
